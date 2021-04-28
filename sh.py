import asyncio
# import util


import os
os.environ["DJANGO_SETTINGS_MODULE"] = "pv-rebalance.settings"
import django
django.setup()
from scraping_events.models import Target, VAAStrategy
# import xlrd
 
# # Give the location of the file
# loc = ("rag.xlsx")
 
# # To open Workbook
# wb = xlrd.open_workbook(loc)
# sheet = wb.sheet_by_index(0)
 
# # For row 0 and column 0
# print(sheet.cell_value(0, 0))
# from io import BytesIO
# wb = load_workbook(filename=BytesIO(input_excel.read()))

# https://www.marsja.se/your-guide-to-reading-excel-xlsx-files-in-python/
# https://stackoverflow.com/questions/20635778/using-openpyxl-to-read-file-from-memory





# from openpyxl import load_workbook
# from io import BytesIO
# import urllib.request

# def load_workbook_from_url(url):
#     file = urllib.request.urlopen(url).read()
#     return load_workbook(filename = BytesIO(file))


# # import openpyxl_extended
# from pathlib import Path

# # xlsx_file = Path('SimData', 'play_data.xlsx')
# book = load_workbook_from_url('https://docs.google.com/spreadsheets/d/1sfN_JTs_-XpM4y3dOd1z6Vtn01qXwrKMbDmuIkKKiLI/edit#gid=65999399') 
# print(book)






# import re, urllib, urllib2

# class Spreadsheet(object):
#     def __init__(self, key):
#         super(Spreadsheet, self).__init__()
#         self.key = key

# class Client(object):
#     def __init__(self, email, password):
#         super(Client, self).__init__()
#         self.email = email
#         self.password = password

#     def _get_auth_token(self, email, password, source, service):
#         url = "https://www.google.com/accounts/ClientLogin"
#         params = {
#             "Email": email, "Passwd": password,
#             "service": service,
#             "accountType": "HOSTED_OR_GOOGLE",
#             "source": source
#         }
#         req = urllib2.Request(url, urllib.urlencode(params))
#         return re.findall(r"Auth=(.*)", urllib2.urlopen(req).read())[0]

#     def get_auth_token(self):
#         source = type(self).__name__
#         return self._get_auth_token(self.email, self.password, source, service="wise")

#     def download(self, spreadsheet, gid=0, format="csv"):
#         url_format = "https://docs.google.com/spreadsheets/d/1sfN_JTs_-XpM4y3dOd1z6Vtn01qXwrKMbDmuIkKKiLI/edit#gid=65999399"
#         headers = {
#             "Authorization": "GoogleLogin auth=" + self.get_auth_token(),
#             "GData-Version": "3.0"
#         }
#         req = urllib2.Request(url_format % (spreadsheet.key, format, gid), headers=headers)
#         return urllib2.urlopen(req)

# if __name__ == "__main__":
#     import getpass
#     import csv

#     email = "" # (your email here)
#     password = getpass.getpass()
#     spreadsheet_id = "" # (spreadsheet id here)

#     # Create client and spreadsheet objects
#     gs = Client(email, password)
#     ss = Spreadsheet(spreadsheet_id)

#     # Request a file-like object containing the spreadsheet's contents
#     csv_file = gs.download(ss)

#     # Parse as CSV and print the rows
#     for row in csv.reader(csv_file):
#         print ", ".join(row)


# import openpyxl
# from pathlib import Path

# # xlsx_file = Path('SimData', 'play_data.xlsx')
# wb_obj = openpyxl.load_workbook('rag.xlsx') 

# # Read the active sheet:
# sheet = wb_obj.active


# print(sheet)

# Read the active sheet:
# sheet = wb_obj.active

# print(sheet)
# print(sheet["C2"].value)




# import pandas as pd

# df = pd.read_csv('https://docs.google.com/spreadsheets/d/1sfN_JTs_-XpM4y3dOd1z6Vtn01qXwrKMbDmuIkKKiLI/edit#gid=65999399',
#                  sep='\t',
#                  parse_dates=[0],
#                  names=['a','b','c','d','e','f'])


# print(df)
# import requests

# csv_url = 'https://docs.google.com/spreadsheets/d/1sfN_JTs_-XpM4y3dOd1z6Vtn01qXwrKMbDmuIkKKiLI/edit#gid=65999399'

# req = requests.get(csv_url)
# url_content = req.content
# csv_file = open('downloaded.csv', 'wb')



# csv_file.write(url_content)
# csv_file.close()



# import csv
# with open('downloaded.csv', 'r') as file:
#     reader = csv.reader(file)
#     for row in reader:
#         print(row)


# import pandas as pd
  
# # read an excel file and convert 
# # into a dataframe object
# df = pd.DataFrame(pd.read_excel("rag.xlsx"))
  
# # show the dataframe
# print(df)

import openpyxl
from pathlib import Path 

xlsx_file = 'rag.xlsx'
wb_obj = openpyxl.load_workbook(xlsx_file, data_only=True) 

# Read the active sheet:
sheet = wb_obj.active


result = []
ticker = "AQ2"
target = "AQ97"
for i in range(1, 8):
	print("-----")
	print(ticker)
	print(target)
	print('-----')
	if sheet[target].value == 1:
		result.append({'ticker': sheet[ticker].value, 'target': sheet[target].value  * 100, 'strategy': 'VAA Strategy'})
	else:
		result.append({'ticker': sheet[ticker].value, 'target': sheet[target].value  * 100 + 0.01 if sheet[target].value != 0 else 0, 'strategy': 'VAA Strategy'})
	get_tick = ticker[1]
	get_tar = target[1]
	ticker = ticker.replace(ticker[1], chr(ord(get_tick) + 1)) 
	target = target.replace(target[1], chr(ord(get_tar) + 1))

print(result)


for target in result:
	VAAStrategy.objects.create(**target)


# for i in 

# AQ  97
# AR  97


# position = ""
# for 
# import pandas as pd
# # import pandas as pd



# # df = pd.read_excel('VAA Strategy.Tiingo.xlsx')
# # print(df)
# file_name = "rag.xlsx"
# xl_file = pd.ExcelFile(file_name)

# dfs = {sheet_name: xl_file.parse(sheet_name) for sheet_name in xl_file.sheet_names}


# from xlrd import open_workbook

# class Arm(object):
#     def __init__(self, id, dsp_name, dsp_code, hub_code, pin_code, pptl):
#         self.id = id
#         self.dsp_name = dsp_name
#         self.dsp_code = dsp_code
#         self.hub_code = hub_code
#         self.pin_code = pin_code
#         self.pptl = pptl

#     def __str__(self):
#         return("Arm object:\n"
#                "  Arm_id = {0}\n"
#                "  DSPName = {1}\n"
#                "  DSPCode = {2}\n"
#                "  HubCode = {3}\n"
#                "  PinCode = {4} \n"
#                "  PPTL = {5}"
#                .format(self.id, self.dsp_name, self.dsp_code,
#                        self.hub_code, self.pin_code, self.pptl))

# wb = open_workbook('VAA Strategy.Tiingo.xlsx')
# for sheet in wb.sheets():
#     number_of_rows = sheet.nrows
#     number_of_columns = sheet.ncols

#     items = []

#     rows = []
#     for row in range(1, number_of_rows):
#         values = []
#         for col in range(number_of_columns):
#             value  = (sheet.cell(row,col).value)
#             try:
#                 value = str(int(value))
#             except ValueError:
#                 pass
#             finally:
#                 values.append(value)
#         item = Arm(*values)
#         items.append(item)

# for item in items:
#     print item
#     print("Accessing one single value (eg. DSPName): {0}".format(item.dsp_name))
#     print

# import io
# import requests

# url = "https://docs.google.com/spreadsheets/d/1sfN_JTs_-XpM4y3dOd1z6Vtn01qXwrKMbDmuIkKKiLI/edit#gid=65999399"
# s = requests.get(url).content
# c = pd.read_csv(io.StringIO(s.decode('utf-8')), error_bad_lines=False)
# print(c)